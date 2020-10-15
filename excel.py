from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import writer
import string
from datetime import datetime
import os, sys
import db


in_alpahbet = ['-', 'а', 'б', 'в', 'г', 'д', 'е', 'ё', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п',
               'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ъ', 'ы', 'ь', 'э', 'ю', 'я',
               'А', 'Б', 'В', 'Г', 'Д', 'Е', 'Ё', 'Ж', 'З', 'И', 'Й', 'К', 'Л', 'М', 'Н', 'О', 'П',
               'Р', 'С', 'Т', 'У', 'Ф', 'Х', 'Ц', 'Ч', 'Ш', 'Щ', 'Ъ', 'Ы', 'Ь', 'Э', 'Ю', 'Я', ' ',
               '1', '2', '3', '4', '5', '6', '7', '8', '9', '0',
               'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r',
               's', 't', 'u', 'v', 'w', 'x', 'y', 'z',
               'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
               'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
               '(', ')'
               ]

out_alpahbet = ['-', 'a', 'b', 'v', 'g', 'd', 'e', 'io', 'zh', 'z', 'i', 'y', 'k', 'l', 'm', 'n', 'o', 'p',
                'r', 's', 't', 'u', 'f', 'h', 'ts', 'ch', 'sh', 'sht', 'a', 'i', 'y', 'e', 'yu', 'ya',
                'A', 'B', 'V', 'G', 'D', 'E', 'Io', 'Zh', 'Z', 'I', 'Y', 'K', 'L', 'M', 'N', 'O', 'P',
                'R', 'S', 'T', 'U', 'F', 'H', 'Ts', 'Ch', 'Sh', 'Sht', 'A', 'I', 'Y', 'e', 'Yu', 'Ya', '_',
                '1', '2', '3', '4', '5', '6', '7', '8', '9', '0',
                'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r',
                's', 't', 'u', 'v', 'w', 'x', 'y', 'z',
                'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                '(', ')'
               ]


def translate_str(data):

    result = ''

    for letter in data:
        if letter in in_alpahbet:
            result = result + str(out_alpahbet[in_alpahbet.index(letter)])
    result = result if result != '' else data
    return result


def products_to_file(data):

    wb = Workbook()
    dest_filename = 'D:\Internal_Project\py_updater_site\products.xlsx'
    ws1 = wb.active
    ws1.title = "Products"

    headers = ['product_id', 'name(ru-ru)', 'name(uk-ua)', 'categories',
               'sku', 'upc', 'ean', 'jan', 'isbn', 'mpn', 'location',
               'quantity', 'model', 'manufacturer', 'image_name', 'shipping',
               'price', 'points', 'date_added', 'date_modified', 'date_available',
               'weight', 'weight_unit', 'length', 'width', 'height', 'length_unit',
               'status', 'tax_class_id', 'description(ru-ru)', 'description(uk-ua)',
               'meta_title(ru-ru)', 'meta_title(uk-ua)', 'meta_description(ru-ru)',
               'meta_description(uk-ua)', 'meta_keywords(ru-ru)', 'meta_keywords(uk-ua)',
               'stock_status_id', 'store_ids', 'layout', 'related_ids', 'tags(ru-ru)',
               'tags(uk-ua)', 'sort_order', 'subtract', 'minimum']

    z = 1
    for header in headers:
        ws1.cell(1, z, header)
        z += 1

    i = 2
    for rows in data:
        ws1.cell(i, 1, rows.get('id'))
        ws1.cell(i, 2, rows.get('data'))
        ws1.cell(i, 3, rows.get('data'))
        ws1.cell(i, 4, rows.get('parent'))
        ws1.cell(i, 5, rows.get('sku'))
        ws1.cell(i, 12, rows.get('quantity') if rows.get('quantity') else '0')
        ws1.cell(i, 13, rows.get('sku'))
        ws1.cell(i, 14, rows.get('manufacture'))
        ws1.cell(i, 16, 'yes')
        ws1.cell(i, 17, rows.get('price_opt'))
        ws1.cell(i, 18, '0')
        ws1.cell(i, 19, datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(i, 20, datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(i, 21, datetime.today().strftime("%Y-%m-%d"))
        ws1.cell(i, 22, '0,00')
        ws1.cell(i, 23, 'kg')
        ws1.cell(i, 24, '0')
        ws1.cell(i, 25, '0')
        ws1.cell(i, 26, '0')
        ws1.cell(i, 27, 'cm')
        ws1.cell(i, 28, 'true')
        ws1.cell(i, 29, '9')
        ws1.cell(i, 32, translate_str(rows.get('data')).lower())
        ws1.cell(i, 33, translate_str(rows.get('data')).lower())
        ws1.cell(i, 38, '5')
        ws1.cell(i, 39, '0')
        ws1.cell(i, 44, '0')
        ws1.cell(i, 45, 'true')
        ws1.cell(i, 46, '1')
        i += 1

    wb.save(dest_filename)


def categories_to_file(data):

    wb = Workbook()

    headers = ['category_id', 'parent_id', 'name(ru-ru)', 'name(uk-ua)',
               'top', 'columns', 'sort_order', 'image_name', 'date_added',
               'date_modified', 'description(ru-ru)', 'description(uk-ua)',
               'meta_title(ru-ru)', 'meta_title(uk-ua)', 'meta_description(ru-ru)',
               'meta_description(uk-ua)', 'meta_keywords(ru-ru)', 'meta_keywords(uk-ua)',
               'store_ids', 'layout', 'status']

    dest_filename = 'D:\Internal_Project\py_updater_site\category.xlsx'

    ws1 = wb.active
    ws1.title = "Categories"

    z = 1
    for header in headers:
        ws1.cell(1, z, header)
        z += 1

    i = 2
    for rows in data:
        ws1.cell(i, 1, rows.get('id'))
        ws1.cell(i, 2, rows.get('parent'))
        ws1.cell(i, 3, rows.get('data'))
        ws1.cell(i, 4, rows.get('data'))
        ws1.cell(i, 5, 'true' if rows.get('parent') == 0 else 'false')
        ws1.cell(i, 6, '4')
        ws1.cell(i, 7, '0')
        ws1.cell(i, 9, datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(i, 10, datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(i, 13, translate_str(rows.get('data')).lower())
        ws1.cell(i, 14, translate_str(rows.get('data')).lower())
        ws1.cell(i, 19, '0')
        ws1.cell(i, 21, 'true')
        i += 1
    wb.save(dest_filename)


def run(path):
    # sql = db.DataBases('db.jaguar-team.com', 'gik_dbuser1', 'Vfuytpbz10#', 'gik_moc_db1', '3357')
    # sql.connect()
    sql_aws = db.DataBases('18.196.35.71', 'user_gik_prod', 'gSC*4DJxspp^&H&6',
                           'db_gik_prod_02', '3306')
    sql_aws.connect()

    wb = load_workbook(path)
    sheet_ranges = wb.active
    other = {}
    level_tree = {}
    id_cat = 1
    id_prod = 1
    item_type = {0: 'category', 1: 'product'}
    categories = []
    products = []

    for x in range(sheet_ranges.max_row):

        if x < 2:
            continue
        print(x)
        array_rows = {}

        current_level = sheet_ranges.row_dimensions[x].outlineLevel
        cell_name = str(sheet_ranges.cell(x, 2).value).strip()
        nds = str(sheet_ranges.cell(x, 4).value).strip()
        sku = str(sheet_ranges.cell(x, 1).value).strip()

        row_type = 0 if sheet_ranges.cell(x, 1).value is None else 1

        if current_level == 0:
            other['manufacture'] = ''
            other['level_0'] = cell_name

        level_tree[current_level] = level_tree.get(((current_level - 1) if current_level > 0 else 0)) if row_type == 1 else id_cat
        if (cell_name == '---                                                                   ' or cell_name == '---' or cell_name == '---+') or current_level == 4:
            level_tree[current_level] = level_tree.get(((current_level - 1) if current_level > 0 else 0))
            other['manufacture'] = cell_name if current_level == 4 else other.get('manufacture')
        else:
            if (other.get('level_0') != '06 - ЗАПЧАСТИ') and (other.get('level_0') != '07 - РАЗНОЕ (УСТАРЕВШЕЕ)') and (other.get('level_0') != '09 - ДРУГОЕ'):
                if nds != '+':
                    price = str(sheet_ranges.cell(x, 7).value)
                    price = "0.00" if price == 'None' or price == '' else price
                    price = str(round((float(price) * 1.16) + 0.5))
                    quantity = sheet_ranges.cell(x, 5).value
                    # print('quantity')
                    # print(quantity)
                    quantity = 0 if quantity == '' else quantity
                    quantity = 0 if quantity == None else quantity
                    # print(quantity)
                    if row_type == 1:
                        query_select_id = "SELECT `product_id` FROM `oc_product` WHERE `sku`=" + str(sku)
                        data = sql_aws.query(query_select_id, '', 'S')
                        if len(data) > 0:
                            query = "UPDATE `oc_product` SET `price`=%s, `quantity`=%s, `status`='1' WHERE `sku`=%s"
                            sql_aws.query(query, (price, quantity, sku), 'U')
                        else:
                            array_rows['id'] = sku if row_type == 1 else id_cat
                            array_rows['level'] = current_level
                            array_rows['data'] = cell_name
                            array_rows['parent'] = 0 if current_level == 0 else level_tree.get(((current_level - 1) if current_level > 0 else 0))
                            array_rows['type'] = row_type
                            array_rows['type_name'] = item_type.get(row_type)
                            array_rows['top'] = 'true' if current_level == 0 else 'false'
                            array_rows['sku'] = sku
                            array_rows['quantity'] = quantity
                            array_rows['price_roz'] = sheet_ranges.cell(x, 6).value
                            array_rows['price_opt'] = price
                            array_rows['manufacture'] = other.get('manufacture')

                    else:
                        array_rows['id'] = sku if row_type == 1 else id_cat
                        array_rows['level'] = current_level
                        array_rows['data'] = cell_name
                        array_rows['parent'] = 0 if current_level == 0 else level_tree.get(((current_level - 1) if current_level > 0 else 0))
                        array_rows['type'] = row_type
                        array_rows['type_name'] = item_type.get(row_type)
                        array_rows['top'] = 'true' if current_level == 0 else 'false'

                    if (len(array_rows)):
                        if row_type == 1:
                            print(array_rows)
                            products.append(array_rows)
                        else:
                            categories.append(array_rows)

                    if row_type == 1:
                        id_prod += 1
                    else:
                        id_cat += 1

                else:
                    query = 'UPDATE `oc_product` SET `status`=%s WHERE `sku`=%s'
                    sql_aws.query(query, (str('0'), str(sku)), 'U')

    categories_to_file(categories)
    products_to_file(products)


if len(sys.argv) != 2:
    print('Usage: {} URL'.format(sys.argv[0]))
    exit(1)

run(sys.argv[1])
