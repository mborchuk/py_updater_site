from openpyxl import load_workbook
from openpyxl import Workbook
import sys


def renamerev(seq, x):
    for i in reversed(range(len(seq))):
        if seq[i] == x:
            return i
    return 0

def rename(seq, x):
    for i in range(len(seq)):
        if seq[i] == x:
            if seq[i - 1] == ' ' and seq[i + 1] == ' ':
                return i
    return 0

def run(path, num_row):
    wb = load_workbook(path)
    sheet_ranges = wb.active

    for x in range(sheet_ranges.max_row):

        if x < 2:
            continue
        first_part = second_part = manufacture = ''
        cell_name = sheet_ranges.cell(x, num_row).value
        first_scobka = renamerev(cell_name, '(')
        if cell_name[first_scobka - 1] != ' ':
            first_scobka = 0
        first_score = rename(cell_name, '-')
        start_wthout_score = end_wthout_score = first_score
        if cell_name[first_score - 1] == ' ':
            start_wthout_score = first_score - 1
        if cell_name[first_score + 1] == ' ':
            end_wthout_score = first_score + 2
        if first_scobka >= 10:
            second_scobka = renamerev(cell_name, ')')
            if second_scobka != len(cell_name):
                manufacture = ' ' + cell_name[first_scobka + 1:second_scobka] + ' '
        if start_wthout_score != first_score and end_wthout_score != first_score:
            first_part = str(cell_name[:start_wthout_score])
            second_part = str(cell_name[end_wthout_score:first_scobka - 1])
        else:
            if first_scobka >= 10:
                first_scobka = first_scobka - 1
            else:
                first_scobka = len(cell_name)
            first_part = str(cell_name[end_wthout_score:first_scobka])
        full = (first_part + manufacture + second_part)
        sheet_ranges.cell(x, num_row).value = full
    wb.save(path)


run(sys.argv[1], 2)
run(sys.argv[1], 3)
